import pandas as pd
import os
from openpyxl import load_workbook




#DO NOT USE -- DESTROYS EXCEL FILES



# Input and output file paths
input_file = 'Input/urbs_intertemporal_2050/2024.xlsx'
output_folder = "Input/urbs_intertemporal_2050"
os.makedirs(output_folder, exist_ok=True)

for year in range(2025, 2051):
    output_file = os.path.join(output_folder, f'{year}.xlsx')

    # Copy the original file to the new year file if it does not exist yet
    if not os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            with pd.ExcelFile(input_file) as xls:
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load workbook and access only the "SupIm" sheet
    wb = load_workbook(output_file)
    if 'SupIm' in wb.sheetnames:
        ws = wb['SupIm']

        # Find the correct columns for 't', 'EU27.WindOff', 'EU27.WindOn', and 'EU27.Hydro'
        headers = {cell.value: cell.column for cell in ws[1]}  # Get column indexes from header row

        # Check if required columns exist
        if all(col in headers for col in ['t', 'EU27.WindOff', 'EU27.WindOn', 'EU27.Hydro']):
            # Update rows based on the value in the 't' column
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                t_cell = row[headers['t'] - 1]  # Access cell in 't' column

                if t_cell.value == 0:
                    row[headers['EU27.WindOff'] - 1].value = 0.0
                    row[headers['EU27.WindOn'] - 1].value = 0.0
                    row[headers['EU27.Hydro'] - 1].value = 0.0
                elif t_cell.value == 1:
                    row[headers['EU27.WindOff'] - 1].value = 0.303
                    row[headers['EU27.WindOn'] - 1].value = 0.48
                    row[headers['EU27.Hydro'] - 1].value = 0.207

            # Save the workbook with only the "SupIm" sheet updated
            wb.save(output_file)
            print(f"Updated 'SupIm' sheet for year {year} in file: {output_file}")
        else:
            print("Error: One or more required columns not found in 'SupIm' sheet.")
    else:
        print("Error: 'SupIm' sheet not found in the workbook.")
